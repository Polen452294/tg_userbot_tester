from __future__ import annotations

import io
import os
import tempfile
from dataclasses import dataclass
from typing import Iterable, Optional
from datetime import datetime

from openpyxl import load_workbook, Workbook


# Варианты названий колонок во входном Excel
INN_HEADERS = {"инн", "inn", "tax_id", "taxid"}
FIO_HEADERS = {"фио", "fio", "full_name", "fullname", "name"}


def _norm_header(s: str) -> str:
    return " ".join((s or "").strip().lower().split())


def find_columns(headers: list[str]) -> tuple[int, int]:
    """
    Возвращает индексы колонок (inn_col, fio_col) (0-based).
    Бросает ValueError если не найдено.
    """
    inn_col = -1
    fio_col = -1
    for idx, h in enumerate(headers):
        nh = _norm_header(h)
        if nh in INN_HEADERS and inn_col == -1:
            inn_col = idx
        if nh in FIO_HEADERS and fio_col == -1:
            fio_col = idx

    if inn_col == -1 or fio_col == -1:
        raise ValueError(
            "Не нашёл колонки. Нужны колонки с названиями ИНН/INN и ФИО/FIO "
            f"(нашёл headers={headers})."
        )
    return inn_col, fio_col


@dataclass
class InputRow:
    row_index: int  # 1-based excel row index
    inn: str
    fio: str


def read_input_xlsx(data: bytes) -> list[InputRow]:
    """
    Читает входной xlsx (bytes), вытаскивает ИНН и ФИО.
    """
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active

    # читаем заголовки из первой строки
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value).strip() if cell.value is not None else "")

    inn_col, fio_col = find_columns(headers)

    rows: list[InputRow] = []
    for r in range(2, ws.max_row + 1):
        inn_val = ws.cell(row=r, column=inn_col + 1).value
        fio_val = ws.cell(row=r, column=fio_col + 1).value

        inn = _normalize_inn(inn_val)
        fio = str(fio_val).strip() if fio_val is not None else ""

        if not inn and not fio:
            continue  # пустая строка

        rows.append(InputRow(row_index=r, inn=inn, fio=fio))

    return rows


def _timestamp() -> str:
    """
    Формат: 2026-02-15_14-32
    """
    return datetime.now().strftime("%Y-%m-%d_%H-%M")


def write_output_xlsx(
    *,
    input_rows: list[InputRow],
    results: list[dict[str, str]],
    ) -> tuple[str, str]:
    wb = Workbook()
    ws = wb.active
    ws.title = "results"

    ws.append(["ИНН", "ФИО", "Телефон", "Email", "Статус"])

    for res in results:
        ws.append([
            res.get("inn", ""),
            res.get("fio", ""),
            res.get("phone", ""),
            res.get("email", ""),
            res.get("status", ""),  # <-- пишем код статуса
        ])

    ts = _timestamp()
    filename = f"output_{ts}.xlsx"

    fd, path = tempfile.mkstemp(prefix="userbot_tester_", suffix=".xlsx")
    os.close(fd)
    wb.save(path)
    return path, filename


def write_pending_xlsx(
    *,
    pending_rows: list[InputRow],
) -> tuple[str, str]:
    """
    Пишет Excel только с ИНН/ФИО для необработанных строк.
    Возвращает (path, filename).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "pending"
    ws.append(["ИНН", "ФИО"])

    for r in pending_rows:
        ws.append([r.inn, r.fio])

    ts = _timestamp()
    filename = f"pending_{ts}.xlsx"

    fd, path = tempfile.mkstemp(prefix="userbot_tester_pending_", suffix=".xlsx")
    os.close(fd)
    wb.save(path)

    return path, filename

def _normalize_inn(value) -> str:
    """
    Убирает .0 если Excel передал число как float.
    """
    if value is None:
        return ""

    # Если это float типа 2222058686.0
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)

    # Если это int
    if isinstance(value, int):
        return str(value)

    # Если строка
    s = str(value).strip()

    # Если строка заканчивается на .0 — убираем
    if s.endswith(".0"):
        try:
            return str(int(float(s)))
        except Exception:
            pass

    return s


def read_input_xlsx(data: bytes) -> list[InputRow]:
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active

    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value).strip() if cell.value is not None else "")

    inn_col, fio_col = find_columns(headers)

    rows: list[InputRow] = []
    for r in range(2, ws.max_row + 1):
        inn_val = ws.cell(row=r, column=inn_col + 1).value
        fio_val = ws.cell(row=r, column=fio_col + 1).value

        inn = _normalize_inn(inn_val)
        fio = str(fio_val).strip() if fio_val is not None else ""

        if not inn and not fio:
            continue

        rows.append(InputRow(row_index=r, inn=inn, fio=fio))

    return rows